import json
import openpyxl
from collections import defaultdict

# Load the Excel file
wb = openpyxl.load_workbook('Weekly Task - Rituva.xlsx')

# Create a mapping of (category, element, subelement, task) to weekly task
weekly_tasks_map = defaultdict(str)

# Process Jan 2026 sheet
print("Processing Jan 2026 sheet...")
ws_jan = wb['Jan 2026']
category = None
for row in ws_jan.iter_rows(min_row=1, values_only=True):
    # Skip empty rows
    if not any(row):
        continue
    
    # Check if this is a category header (first cell has value, others empty)
    if row[0] and not row[1]:
        category = row[0]
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Extract the weekly task data
    if len(row) >= 5 and row[2] is not None and row[4] is not None:  # Task and Weekly Task
        element = row[0]
        subelement = row[1]
        task = row[2]
        weekly_task = row[4]
        
        # Create a composite key
        key = (category, element, subelement, task)
        weekly_tasks_map[key] = weekly_task
        print(f"✓ Jan: {category} - {subelement} - Weekly task added")

# Process Feb 2026 sheet
print("\nProcessing Feb 2026 sheet...")
ws_feb = wb['Feb 2026']
category = None
for row in ws_feb.iter_rows(min_row=1, values_only=True):
    # Skip empty rows
    if not any(row):
        continue
    
    # Check if this is a category header
    if row[0] and not row[1]:
        category = row[0]
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Extract the weekly task data
    if len(row) >= 5 and row[2] is not None and row[4] is not None:
        element = row[0]
        subelement = row[1]
        task = row[2]
        weekly_task = row[4]
        
        # Create a composite key
        key = (category, element, subelement, task)
        weekly_tasks_map[key] = weekly_task
        print(f"✓ Feb: {category} - {subelement} - Weekly task added")

print(f"\n{'='*70}")
print(f"Total weekly tasks extracted: {len(weekly_tasks_map)}")
print(f"{'='*70}\n")

# Now load and update the digital-marketing-data.json
with open('public/digital-marketing-data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Update entries
updated_count = 0
matched_count = 0
not_matched = []

for entry in data:
    if entry.get('month') not in ['January', 'February']:
        continue
    
    # Create a key from the entry
    category = entry.get('category', '')
    element = entry.get('element', '')
    subelement = entry.get('subElement', '')
    task = entry.get('task', '')
    
    key = (category, element, subelement, task)
    
    if key in weekly_tasks_map:
        entry['weeklyTask'] = weekly_tasks_map[key]
        updated_count += 1
        matched_count += 1
        print(f"✓ MATCHED: [{entry['month']}] {category} - {subelement}")
    else:
        # Try partial matching (category + subelement)
        partial_key = (category, element, subelement)
        for full_key, weekly_task in weekly_tasks_map.items():
            if full_key[0] == category and full_key[1] == element and full_key[2] == subelement:
                entry['weeklyTask'] = weekly_task
                updated_count += 1
                matched_count += 1
                print(f"✓ PARTIAL MATCHED: [{entry['month']}] {category} - {subelement}")
                break

# Save the updated data
with open('public/digital-marketing-data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"\n{'='*70}")
print(f"✅ INTEGRATION COMPLETED")
print(f"{'='*70}")
print(f"✓ Total entries updated with weekly tasks: {updated_count}")
print(f"✓ Total matched: {matched_count}")
print(f"{'='*70}")
print(f"✅ File updated successfully!")

