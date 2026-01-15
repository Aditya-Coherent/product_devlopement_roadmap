import json
import openpyxl

print("="*80)
print("PRECISE WEEKLY TASK MAPPING - Starting from scratch")
print("="*80)

# Step 1: Load JSON and clear ALL existing weeklyTask fields
with open('public/digital-marketing-data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

cleared_count = 0
for entry in data:
    if 'weeklyTask' in entry:
        del entry['weeklyTask']
        cleared_count += 1

print(f"✓ Cleared {cleared_count} existing weeklyTask fields")

# Step 2: Load Excel
wb = openpyxl.load_workbook('Weekly Task - Rituva.xlsx')

# Step 3: Create PRECISE mappings from Excel
# Key = (task_text_normalized) -> (category, subelement, weekly_task)
# We use task as the primary key since it's most unique

def normalize(text):
    """Normalize text for comparison"""
    if text is None:
        return ""
    return str(text).strip().lower()

january_mappings = {}  # task_normalized -> (subelement, weekly_task)
february_mappings = {}

# Process January sheet
print("\n" + "="*80)
print("Processing January 2026 sheet...")
print("="*80)
ws_jan = wb['Jan 2026']
current_category = None
current_element = None
current_subelement = None

for row in ws_jan.iter_rows(min_row=1, values_only=True):
    if not any(row):
        continue
    
    # Category header (first column only, no other values)
    if row[0] and not row[1] and not row[2]:
        current_category = str(row[0]).strip()
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Data row
    element = row[0] if row[0] else current_element
    subelement = row[1] if row[1] else current_subelement
    task = row[2]
    monthly = row[3]
    weekly = row[4]
    
    # Update current values
    if row[0]:
        current_element = row[0]
    if row[1]:
        current_subelement = row[1]
    
    # Only add if task exists and weekly task exists
    if task and weekly:
        task_norm = normalize(task)
        subelement_norm = normalize(subelement)
        
        key = (task_norm, subelement_norm)
        january_mappings[key] = {
            'category': current_category,
            'subelement': subelement,
            'task': str(task),
            'weeklyTask': str(weekly)
        }
        print(f"✓ JAN: {str(subelement)[:30]} | {str(task)[:40]}...")

print(f"\nTotal January mappings: {len(january_mappings)}")

# Process February sheet
print("\n" + "="*80)
print("Processing February 2026 sheet...")
print("="*80)
ws_feb = wb['Feb 2026']
current_category = None
current_element = None
current_subelement = None

for row in ws_feb.iter_rows(min_row=1, values_only=True):
    if not any(row):
        continue
    
    # Category header
    if row[0] and not row[1] and not row[2]:
        current_category = str(row[0]).strip()
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Data row
    element = row[0] if row[0] else current_element
    subelement = row[1] if row[1] else current_subelement
    task = row[2]
    monthly = row[3]
    weekly = row[4]
    
    # Update current values
    if row[0]:
        current_element = row[0]
    if row[1]:
        current_subelement = row[1]
    
    # Only add if task exists and weekly task exists
    if task and weekly:
        task_norm = normalize(task)
        subelement_norm = normalize(subelement)
        
        key = (task_norm, subelement_norm)
        february_mappings[key] = {
            'category': current_category,
            'subelement': subelement,
            'task': str(task),
            'weeklyTask': str(weekly)
        }
        print(f"✓ FEB: {str(subelement)[:30]} | {str(task)[:40]}...")

print(f"\nTotal February mappings: {len(february_mappings)}")

# Step 4: Apply mappings to JSON data - STRICTLY by month
print("\n" + "="*80)
print("Applying mappings to JSON...")
print("="*80)

january_updated = 0
february_updated = 0
january_not_found = []
february_not_found = []

for entry in data:
    month = entry.get('month')
    task = entry.get('task', '')
    subelement = entry.get('subElement', '')
    
    task_norm = normalize(task)
    subelement_norm = normalize(subelement)
    key = (task_norm, subelement_norm)
    
    if month == 'January':
        if key in january_mappings:
            entry['weeklyTask'] = january_mappings[key]['weeklyTask']
            january_updated += 1
            print(f"✓ [JAN] Matched: {subelement[:40]}...")
        else:
            # Also try matching by task only if subelement match fails
            matched = False
            for map_key, map_value in january_mappings.items():
                if map_key[0] == task_norm:
                    entry['weeklyTask'] = map_value['weeklyTask']
                    january_updated += 1
                    print(f"✓ [JAN] Task-matched: {subelement[:40]}...")
                    matched = True
                    break
            if not matched and task_norm:
                january_not_found.append(f"{subelement}: {task[:50]}")
                
    elif month == 'February':
        if key in february_mappings:
            entry['weeklyTask'] = february_mappings[key]['weeklyTask']
            february_updated += 1
            print(f"✓ [FEB] Matched: {subelement[:40]}...")
        else:
            # Also try matching by task only
            matched = False
            for map_key, map_value in february_mappings.items():
                if map_key[0] == task_norm:
                    entry['weeklyTask'] = map_value['weeklyTask']
                    february_updated += 1
                    print(f"✓ [FEB] Task-matched: {subelement[:40]}...")
                    matched = True
                    break
            if not matched and task_norm:
                february_not_found.append(f"{subelement}: {task[:50]}")

# Step 5: Save
with open('public/digital-marketing-data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("\n" + "="*80)
print("RESULTS")
print("="*80)
print(f"✓ January entries updated: {january_updated}")
print(f"✓ February entries updated: {february_updated}")
print(f"✓ Total updated: {january_updated + february_updated}")

if january_not_found:
    print(f"\n⚠ January entries without weekly tasks ({len(january_not_found)}):")
    for item in january_not_found[:10]:
        print(f"  - {item}")
    if len(january_not_found) > 10:
        print(f"  ... and {len(january_not_found) - 10} more")

if february_not_found:
    print(f"\n⚠ February entries without weekly tasks ({len(february_not_found)}):")
    for item in february_not_found[:10]:
        print(f"  - {item}")
    if len(february_not_found) > 10:
        print(f"  ... and {len(february_not_found) - 10} more")

print("\n" + "="*80)
print("✅ File updated successfully!")
print("="*80)

