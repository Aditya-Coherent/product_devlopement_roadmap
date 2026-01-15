import json
import openpyxl
from collections import defaultdict

# Read the current JSON file
with open('public/digital-marketing-data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Load the Excel file
wb = openpyxl.load_workbook('Weekly Task - Rituva.xlsx')

# Create separate mappings for each month
january_weekly_tasks = defaultdict(str)
february_weekly_tasks = defaultdict(str)

# Process Jan 2026 sheet
print("Processing Jan 2026 sheet...")
ws_jan = wb['Jan 2026']
category = None
for row in ws_jan.iter_rows(min_row=1, values_only=True):
    # Skip empty rows
    if not any(row):
        continue
    
    # Check if this is a category header
    if row[0] and not row[1]:
        category = row[0]
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Extract the weekly task data
    if len(row) >= 5 and row[2] is not None and row[4] is not None:
        element = row[0]
        subelement = row[1]
        task = row[2]
        weekly_task = row[4]
        
        # Create a composite key
        key = (category, element, subelement, task)
        january_weekly_tasks[key] = weekly_task
        print(f"✓ Jan: {str(subelement)[:30]} - {str(task)[:50]}...")

# Process Feb 2026 sheet
print("\nProcessing Feb 2026 sheet...")
ws_feb = wb['Feb 2026']
category = None
for row in ws_feb.iter_rows(min_row=1, values_only=True):
    # Skip empty rows
    if not any(row):
        continue
    
    # Check if this is a category header
    if row[0] and not row[1]:
        category = row[0]
        continue
    
    # Skip header row
    if row[0] == 'Element':
        continue
    
    # Extract the weekly task data
    if len(row) >= 5 and row[2] is not None and row[4] is not None:
        element = row[0]
        subelement = row[1]
        task = row[2]
        weekly_task = row[4]
        
        # Create a composite key
        key = (category, element, subelement, task)
        february_weekly_tasks[key] = weekly_task
        print(f"✓ Feb: {str(subelement)[:30]} - {str(task)[:50]}...")

print(f"\n{'='*70}")
print(f"Total January weekly tasks: {len(january_weekly_tasks)}")
print(f"Total February weekly tasks: {len(february_weekly_tasks)}")
print(f"{'='*70}\n")

# Now update the JSON data with correct month-specific tasks
january_updated = 0
february_updated = 0

for entry in data:
    if entry.get('month') == 'January':
        category = entry.get('category', '')
        element = entry.get('element', '')
        subelement = entry.get('subElement', '')
        task = entry.get('task', '')
        
        key = (category, element, subelement, task)
        
        if key in january_weekly_tasks:
            entry['weeklyTask'] = january_weekly_tasks[key]
            january_updated += 1
            print(f"✓ UPDATED [JAN]: {str(subelement)[:50]}")
    
    elif entry.get('month') == 'February':
        category = entry.get('category', '')
        element = entry.get('element', '')
        subelement = entry.get('subElement', '')
        task = entry.get('task', '')
        
        key = (category, element, subelement, task)
        
        if key in february_weekly_tasks:
            entry['weeklyTask'] = february_weekly_tasks[key]
            february_updated += 1
            print(f"✓ UPDATED [FEB]: {str(subelement)[:50]}")

# Save the updated data
with open('public/digital-marketing-data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"\n{'='*70}")
print(f"✅ CORRECTION COMPLETED")
print(f"{'='*70}")
print(f"✓ January entries updated: {january_updated}")
print(f"✓ February entries updated: {february_updated}")
print(f"✓ Total entries updated: {january_updated + february_updated}")
print(f"{'='*70}")
print(f"✅ File corrected successfully!")

