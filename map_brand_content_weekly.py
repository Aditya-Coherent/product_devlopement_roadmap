import json
import openpyxl

print("="*80)
print("MAPPING Brand Positioning & Content Marketing - Weekly Tasks")
print("="*80)

# Load JSON
with open('public/digital-marketing-data.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Load Excel
wb = openpyxl.load_workbook('Weekly Task - Rituva.xlsx')

def normalize(text):
    if text is None:
        return ""
    return str(text).strip().lower()

# Extract weekly tasks from Excel for these specific categories
# Structure: (month, task_identifier) -> weekly_task

january_brand_weekly = {}
january_content_weekly = {}
february_brand_weekly = {}
february_content_weekly = {}

# Process January
print("\nProcessing January Excel...")
ws_jan = wb['Jan 2026']
current_category = None
current_element = None

for row in ws_jan.iter_rows(min_row=1, values_only=True):
    if not any(row):
        continue
    
    # Category header
    if row[0] and not row[1] and not row[2]:
        current_category = str(row[0]).strip()
        continue
    
    # Skip header
    if row[0] == 'Element':
        continue
    
    # Update element if present
    if row[0]:
        current_element = str(row[0]).strip()
    
    subelement = row[1]
    task = row[2]
    weekly = row[4]
    
    if current_category == 'Brand Positioning & Thought Leadership':
        # For Brand Positioning, the Excel subelement matches JSON task
        if subelement and weekly:
            key = normalize(subelement)
            january_brand_weekly[key] = str(weekly)
            print(f"✓ JAN Brand: {subelement} -> {str(weekly)[:50]}...")
        # Also check task for seasonal items
        if task and weekly and current_element:
            task_key = normalize(task)
            january_brand_weekly[task_key] = str(weekly)
            
    elif current_category == 'Content Marketing & Strategy':
        # For Content Marketing, the Excel subelement matches JSON task
        if subelement and weekly:
            key = normalize(subelement)
            january_content_weekly[key] = str(weekly)
            print(f"✓ JAN Content: {subelement} -> {str(weekly)[:50]}...")

# Process February
print("\nProcessing February Excel...")
ws_feb = wb['Feb 2026']
current_category = None
current_element = None

for row in ws_feb.iter_rows(min_row=1, values_only=True):
    if not any(row):
        continue
    
    # Category header
    if row[0] and not row[1] and not row[2]:
        current_category = str(row[0]).strip()
        continue
    
    # Skip header
    if row[0] == 'Element':
        continue
    
    # Update element if present
    if row[0]:
        current_element = str(row[0]).strip()
    
    subelement = row[1]
    task = row[2]
    weekly = row[4]
    
    if current_category == 'Brand Positioning & Thought Leadership':
        if subelement and weekly:
            key = normalize(subelement)
            february_brand_weekly[key] = str(weekly)
            print(f"✓ FEB Brand: {subelement} -> {str(weekly)[:50]}...")
        if task and weekly and current_element:
            task_key = normalize(task)
            february_brand_weekly[task_key] = str(weekly)
            
    elif current_category == 'Content Marketing & Strategy':
        if subelement and weekly:
            key = normalize(subelement)
            february_content_weekly[key] = str(weekly)
            print(f"✓ FEB Content: {subelement} -> {str(weekly)[:50]}...")

print(f"\nJanuary Brand mappings: {len(january_brand_weekly)}")
print(f"January Content mappings: {len(january_content_weekly)}")
print(f"February Brand mappings: {len(february_brand_weekly)}")
print(f"February Content mappings: {len(february_content_weekly)}")

# Now apply to JSON
print("\n" + "="*80)
print("Applying mappings to JSON...")
print("="*80)

jan_brand_updated = 0
jan_content_updated = 0
feb_brand_updated = 0
feb_content_updated = 0

for entry in data:
    month = entry.get('month')
    category = entry.get('category')
    task = entry.get('task', '')
    subelement = entry.get('subElement', '')
    
    task_norm = normalize(task)
    subelement_norm = normalize(subelement)
    
    if month == 'January':
        if category == 'Brand Positioning & Thought Leadership':
            # Try matching by task name
            if task_norm in january_brand_weekly:
                entry['weeklyTask'] = january_brand_weekly[task_norm]
                jan_brand_updated += 1
                print(f"✓ [JAN Brand] {task[:40]} -> matched")
            # Also check for partial match
            else:
                for key, weekly in january_brand_weekly.items():
                    if key in task_norm or task_norm in key:
                        entry['weeklyTask'] = weekly
                        jan_brand_updated += 1
                        print(f"✓ [JAN Brand] {task[:40]} -> partial matched")
                        break
                        
        elif category == 'Content Marketing & Strategy':
            # Try matching by task name (which is like "I. Website Content Overhaul...")
            if task_norm in january_content_weekly:
                entry['weeklyTask'] = january_content_weekly[task_norm]
                jan_content_updated += 1
                print(f"✓ [JAN Content] {task[:40]} -> matched")
            else:
                for key, weekly in january_content_weekly.items():
                    if key in task_norm or task_norm in key:
                        entry['weeklyTask'] = weekly
                        jan_content_updated += 1
                        print(f"✓ [JAN Content] {task[:40]} -> partial matched")
                        break
                        
    elif month == 'February':
        if category == 'Brand Positioning & Thought Leadership':
            if task_norm in february_brand_weekly:
                entry['weeklyTask'] = february_brand_weekly[task_norm]
                feb_brand_updated += 1
                print(f"✓ [FEB Brand] {task[:40]} -> matched")
            else:
                for key, weekly in february_brand_weekly.items():
                    if key in task_norm or task_norm in key:
                        entry['weeklyTask'] = weekly
                        feb_brand_updated += 1
                        print(f"✓ [FEB Brand] {task[:40]} -> partial matched")
                        break
                        
        elif category == 'Content Marketing & Strategy':
            if task_norm in february_content_weekly:
                entry['weeklyTask'] = february_content_weekly[task_norm]
                feb_content_updated += 1
                print(f"✓ [FEB Content] {task[:40]} -> matched")
            else:
                for key, weekly in february_content_weekly.items():
                    if key in task_norm or task_norm in key:
                        entry['weeklyTask'] = weekly
                        feb_content_updated += 1
                        print(f"✓ [FEB Content] {task[:40]} -> partial matched")
                        break

# Save
with open('public/digital-marketing-data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print("\n" + "="*80)
print("RESULTS")
print("="*80)
print(f"✓ January Brand Positioning updated: {jan_brand_updated}")
print(f"✓ January Content Marketing updated: {jan_content_updated}")
print(f"✓ February Brand Positioning updated: {feb_brand_updated}")
print(f"✓ February Content Marketing updated: {feb_content_updated}")
print(f"✓ Total updated: {jan_brand_updated + jan_content_updated + feb_brand_updated + feb_content_updated}")
print("="*80)
print("✅ File updated successfully!")

